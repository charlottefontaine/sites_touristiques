"""
corpus_visualisation.py

Builds Excel heatmaps for:
- top 10 most frequent words per city (global + by zone),
- LDA topic proportions per city (global + by zone).

Inputs
------
data/processed/df_freq_terms.csv
    Document-term matrix with a 'city' column.

Outputs
-------
data/text_analysis/heatmaps/heatmap_words_{GLOBAL|SEA|NOSEA|NORTH|SOUTH}.xlsx
data/text_analysis/heatmaps/heatmap_topics_{GLOBAL|SEA|NOSEA|NORTH|SOUTH}.xlsx
"""

import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from parameters import SEA_CITIES, NOSEA_CITIES, NORTH_CITIES, SOUTH_CITIES

BASE_DIR = "data"
PROCESSED_DIR = os.path.join(BASE_DIR, "processed")
HEATMAP_DIR = os.path.join(BASE_DIR, "text_analysis", "heatmaps")

os.makedirs(HEATMAP_DIR, exist_ok=True)


def get_zone(city: str) -> str:
    if city in SEA_CITIES:
        return "SEA"
    if city in NOSEA_CITIES:
        return "NOSEA"
    if city in NORTH_CITIES:
        return "NORTH"
    if city in SOUTH_CITIES:
        return "SOUTH"
    return "Other"


def save_heatmap_excel(df: pd.DataFrame, output_path: str) -> None:
    """
    Apply a simple color scale on numeric cells in an existing Excel file.
    """
    wb = load_workbook(output_path)
    ws = wb.active

    df_numeric = df.select_dtypes(include=[np.number])
    if df_numeric.empty:
        print("No numeric data to color in heatmap.")
        return

    max_val = df_numeric.to_numpy().max()

    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                intensity = int(255 * cell.value / max_val) if max_val else 0
                color = f"FF{255 - intensity:02X}{255 - intensity:02X}"
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

    wb.save(output_path)


def main():
    # Load global df_freq_terms
    freq_path = os.path.join(PROCESSED_DIR, "df_freq_terms.csv")
    if not os.path.exists(freq_path):
        raise FileNotFoundError(
            f"{freq_path} not found. Run corpus_cleaning.py first."
        )

    df = pd.read_csv(freq_path)
    if "city" not in df.columns:
        raise ValueError("Expected 'city' column in df_freq_terms.csv")

    df["Zone"] = df["city"].apply(get_zone)
    terms = [c for c in df.columns if c not in ["city", "Zone"]]
    df[terms] = df[terms].apply(pd.to_numeric, errors="coerce").fillna(0)

    # -----------------------------
    # PART 1 : Word heatmaps
    # -----------------------------
    def word_heatmap(data: pd.DataFrame, name: str):
        freq = data[terms].sum().sort_values(ascending=False).head(10)
        df_hm = data.groupby("city")[freq.index].sum()
        path = os.path.join(HEATMAP_DIR, f"heatmap_words_{name}.xlsx")
        df_hm.to_excel(path)
        save_heatmap_excel(df_hm, path)
        print(f"Word heatmap saved to {path}")

    # Global
    word_heatmap(df, "GLOBAL")
    # By zone
    for zone in ["SEA", "NOSEA", "NORTH", "SOUTH"]:
        word_heatmap(df[df["Zone"] == zone], zone)

if __name__ == "__main__":
    main()