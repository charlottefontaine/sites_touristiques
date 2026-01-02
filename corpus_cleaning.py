from Utils import *
from parameters import *

import pandas as pd
import numpy as np
from sklearn.feature_extraction.text import TfidfTransformer
from sklearn.decomposition import LatentDirichletAllocation
import os
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ---------------------------------
# Part 1 . Term-document matrix
# ---------------------------------
def main(dict_tokens):
    cities = set(key[0] for key in dict_tokens.keys())
    print(f"Cities present: {sorted(cities)}")
    print(f"Total entries: {len(dict_tokens)}")

    df_tdm = build_term_document_matrix(dict_tokens)
    print(f"Term-document matrix shape: {df_tdm.shape}")

    df_tdm = clean_term_document_matrix(df_tdm, None, keep_numbers=False)
    df_tdm = normalize_social_media_terms(df_tdm)
    df_tdm = remove_nltk_stopwords(df_tdm)
    df_tdm = lemmatize_matrix_nltk(df_tdm)
    df_tdm = remove_project_stopwords(df_tdm, project_stopwords)
    df_tdm = remove_miniwords(df_tdm, min_length=3)

    df_freq_terms = filter_terms_by_frequency(df_tdm)
    print(f"Final shape after filtering: {df_freq_terms.shape}")

    # Ajout colonne city
    df_freq_terms = df_freq_terms.copy()
    df_freq_terms["city"] = [t[0] for t in df_freq_terms.index]
    cols = ["city"] + [c for c in df_freq_terms.columns if c != "city"]
    df_freq_terms = df_freq_terms[cols]

    os.makedirs("data/processed", exist_ok=True)
    return df_freq_terms

# ---------------------------------
# Part 2 . Heatmap Excel classique
# ---------------------------------
def save_heatmap_excel(df, output_path, min_count=4, max_words=100):
    df_numeric = df.drop(columns=["city"]).apply(pd.to_numeric, errors="coerce").fillna(0)
    corpus_freq = df_numeric.sum(axis=0)
    corpus_freq = corpus_freq[corpus_freq >= min_count]
    top_terms = corpus_freq.sort_values(ascending=False).head(max_words).index
    df_numeric = df_numeric[top_terms]
    df_to_save = pd.concat([df["city"], df_numeric], axis=1)
    df_to_save.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    max_val = df_numeric.to_numpy().max()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=2, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                intensity = int(255 * cell.value / max_val) if max_val != 0 else 0
                red = 255
                green = 255 - intensity
                blue = 255 - intensity
                cell.fill = PatternFill(start_color=f"{red:02X}{green:02X}{blue:02X}",
                                        end_color=f"{red:02X}{green:02X}{blue:02X}",
                                        fill_type="solid")
    wb.save(output_path)

# ---------------------------------
# Part 3 . Nouveau heatmap LDA global
# ---------------------------------
def save_topic_heatmap_excel(df_topics, zone_map, output_path):
    """
    df_topics: DataFrame (cities x topics)
    zone_map: dict {city: zone_name}
    """
    df_topics = df_topics.copy()
    df_topics["Zone"] = df_topics.index.map(zone_map)
    df_topics = df_topics.sort_values("Zone")
    df_topics.to_excel(output_path, index=True)

    # Heatmap couleur par intensité
    wb = load_workbook(output_path)
    ws = wb.active
    max_val = df_topics.drop(columns=["Zone"]).to_numpy().max()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=2, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                intensity = int(255 * cell.value / max_val) if max_val != 0 else 0
                red = 255
                green = 255 - intensity
                blue = 255 - intensity
                cell.fill = PatternFill(start_color=f"{red:02X}{green:02X}{blue:02X}",
                                        end_color=f"{red:02X}{green:02X}{blue:02X}",
                                        fill_type="solid")
    wb.save(output_path)

# ---------------------------------
# Zones
# ---------------------------------
SEA_CITIES = {"Barcelona", "Lisbon", "Copenhagen", "Ostend","Valencia"}
NOSEA_CITIES = {"Rome", "Manchester", "Cologne" , "Amsterdam","Bruges"}
SOUTH_CITIES = {"Barcelona", "Lisbon", "Rome", "Valencia"}
NORTH_CITIES = {"Amsterdam", "Copenhagen", "Manchester", "Cologne", "Ostend", "Bruges"}

# ---------------------------------
# Main processing
# ---------------------------------
corpus_files = {
    "sea": "data/processed/corpus_sea.json",
    "nosea": "data/processed/corpus_nosea.json",
    "north": "data/processed/corpus_north.json",
    "south": "data/processed/corpus_south.json"
}

all_freq_terms = []

for zone, path in corpus_files.items():
    dict_tokens = tokenize_json_by_city_url(path)
    df_zone = main(dict_tokens)
    all_freq_terms.append(df_zone)

# Concaténer toutes les villes dans un seul DataFrame
df_all_cities = pd.concat(all_freq_terms, ignore_index=True)

# Sauvegarde TDM global
df_all_cities.to_csv("data/processed/df_freq_terms_all.csv", index=False)
save_heatmap_excel(df_all_cities, "data/processed/heatmap_all_zones.xlsx")

# -----------------------------
# LDA pour topic par ville
# -----------------------------
terms = [c for c in df_all_cities.columns if c != "city"]
tdm = df_all_cities[terms].fillna(0).values
cities = df_all_cities["city"].values

lda_model = LatentDirichletAllocation(n_components=6, max_iter=100, random_state=42)
doc_topic = lda_model.fit_transform(tdm)

# Agrégation par ville
df_city_topics = pd.DataFrame(doc_topic, index=cities, columns=[f"Topic_{i}" for i in range(6)])
df_city_topics = df_city_topics.groupby(df_city_topics.index).mean()
df_city_topics = df_city_topics.div(df_city_topics.sum(axis=1), axis=0)

# Mapping zones
zone_map = {}
for city in df_city_topics.index:
    if city in SEA_CITIES:
        zone_map[city] = "SEA"
    elif city in NOSEA_CITIES:
        zone_map[city] = "NOSEA"
    elif city in SOUTH_CITIES:
        zone_map[city] = "SOUTH"
    elif city in NORTH_CITIES:
        zone_map[city] = "NORTH"
    else:
        zone_map[city] = "Other"

# Sauvegarde heatmap LDA global
save_topic_heatmap_excel(df_city_topics, zone_map, "data/processed/heatmap_topic_all_zones.xlsx")
print("Global topic heatmap saved: data/processed/heatmap_topic_all_zones.xlsx")
